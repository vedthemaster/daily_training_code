import openpyxl


class HomePageData:
    # Create Class Variable to Access Data It is Dictionary only Key value Pair
    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

    @staticmethod  # Used for not to create class while using
    def getTestData(test_case_name):
        Dict = {}
        # book = openpyxl.load_workbook("C:\\Users\\Owner\\Documents\\PythonDemo.xlsx")
        book = openpyxl.load_workbook("C:\\Users\\vedpa\\OneDrive\\Desktop\\Training Folder\\Python with Selenium\\PythonSelFramework\\pythondemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]  # To wrap Dictionary for Data Provider

